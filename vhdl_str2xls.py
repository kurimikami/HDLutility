import re
import openpyxl
from openpyxl.styles import PatternFill
import argparse
import os

def parse_vhdl_file(vhdl_file_path):
    with open(vhdl_file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # コメントを無視するため、すべてのコメントを除去
    content = re.sub(r'--.*', '', content)

    # 最上位階層のentity名を取得
    entity_name_match = re.search(r'entity\s+(\w+)\s+is', content)
    top_entity_name = entity_name_match.group(1) if entity_name_match else None

    # Port宣言を抽出し、宣言されているポート名と方向を収集
    declared_ports = {}
    port_declaration_match = re.search(r'port\s*\((.*?)\)\s*;', content, re.DOTALL)
    if port_declaration_match:
        ports = port_declaration_match.group(1)
        for port_name, port_direction in re.findall(r'(\w+)\s*:\s*(\w+)', ports):
            declared_ports[port_name] = port_direction

    # インスタンスのポートマッピングを抽出
    instances = {}
    inst_matches = re.findall(r'(\w+)\s*:\s*(?:entity|component)\s+([\w\.]+)\s*port\s*map\s*\(([^;]+)\);', content, re.DOTALL)
    for inst_name, entity_name, port_map in inst_matches:
        if entity_name not in instances:
            instances[entity_name] = {}
        ports = re.findall(r'(\w+)\s*=>\s*(\w+)', port_map)
        for port_name, net_name in ports:
            if net_name not in instances[entity_name]:
                instances[entity_name][net_name] = {}
            if inst_name in instances[entity_name][net_name]:
                instances[entity_name][net_name][inst_name] += ',' + port_name
            else:
                instances[entity_name][net_name][inst_name] = port_name

    return top_entity_name, declared_ports, instances

def generate_excel(output_file_path, top_entity_name, declared_ports, instances):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = top_entity_name if top_entity_name else "TopEntity"

    # ネット名とインスタンス名の取得
    all_nets = set()
    all_inst_names = set()

    for entity_data in instances.values():
        all_nets.update(entity_data.keys())
        all_inst_names.update({inst for net in entity_data.values() for inst in net.keys()})

    nets = sorted(all_nets)
    inst_names = sorted(all_inst_names)

    # 1行目にインスタンス名を書き込む（C1から）
    for col, inst_name in enumerate(inst_names, start=3):
        ws.cell(row=1, column=col, value=inst_name)

    # 1列目にネット名を書き込む（B2から）
    for row, net_name in enumerate(nets, start=2):
        net_name_orig = net_name.strip("()")
        direction = declared_ports.get(net_name_orig, "")

        if net_name_orig not in declared_ports:
            cell = ws.cell(row=row, column=2, value=f"({net_name})")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        else:
            ws.cell(row=row, column=1, value=direction)
            ws.cell(row=row, column=2, value=net_name_orig)

        for col, inst_name in enumerate(inst_names, start=3):
            port_name = ""
            for entity_data in instances.values():
                if net_name_orig in entity_data and inst_name in entity_data[net_name_orig]:
                    port_name = entity_data[net_name_orig][inst_name]
            ws.cell(row=row, column=col, value=port_name)

    wb.save(output_file_path)
    print(f"{output_file_path} が生成されました。")

def main():
    parser = argparse.ArgumentParser(description="VHDLファイルからネットとインスタンスの接続情報をExcelに出力するツール")
    parser.add_argument("input_file", help="入力VHDLファイルのパス")
    parser.add_argument("output_file", help="出力Excelファイルのパス")
    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"エラー: 指定された入力ファイル '{args.input_file}' が存在しません。")
        return

    try:
        top_entity_name, declared_ports, instances = parse_vhdl_file(args.input_file)
        generate_excel(args.output_file, top_entity_name, declared_ports, instances)
    except Exception as e:
        print(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    main()
